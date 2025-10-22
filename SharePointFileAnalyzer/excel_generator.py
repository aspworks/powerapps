"""
Excel file generator for analysis results
"""
from typing import List, Dict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Generator for creating Excel reports"""

    def __init__(self, output_filename: str = 'sharepoint_file_analysis.xlsx'):
        """
        Initialize Excel generator

        Args:
            output_filename: Name of the output Excel file
        """
        self.output_filename = output_filename
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "File Analysis"

    def create_report(self, analysis_results: List[Dict], sharepoint_url: str, folder_path: str):
        """
        Create Excel report from analysis results

        Args:
            analysis_results: List of analysis result dictionaries
            sharepoint_url: SharePoint site URL
            folder_path: Folder path that was analyzed
        """
        # Add title
        self.ws['A1'] = "SharePoint File Analysis Report"
        self.ws['A1'].font = Font(size=16, bold=True)
        self.ws.merge_cells('A1:E1')

        # Add metadata
        self.ws['A2'] = f"SharePoint Site: {sharepoint_url}"
        self.ws['A3'] = f"Folder Path: {folder_path}"
        self.ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        self.ws['A5'] = f"Total Files Analyzed: {len(analysis_results)}"

        # Add headers
        headers = ['#', 'Filename', 'Title', 'Summary', 'File Size (KB)', 'Last Modified']
        header_row = 7

        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add data
        for idx, result in enumerate(analysis_results, 1):
            row = header_row + idx
            self.ws.cell(row=row, column=1, value=idx)
            self.ws.cell(row=row, column=2, value=result.get('filename', ''))
            self.ws.cell(row=row, column=3, value=result.get('title', ''))
            self.ws.cell(row=row, column=4, value=result.get('summary', ''))
            self.ws.cell(row=row, column=5, value=round(result.get('size', 0) / 1024, 2))
            self.ws.cell(row=row, column=6, value=result.get('time_modified', ''))

            # Enable text wrapping for summary column
            self.ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')

        # Adjust column widths
        self.ws.column_dimensions['A'].width = 5
        self.ws.column_dimensions['B'].width = 30
        self.ws.column_dimensions['C'].width = 35
        self.ws.column_dimensions['D'].width = 60
        self.ws.column_dimensions['E'].width = 15
        self.ws.column_dimensions['F'].width = 20

        # Set row heights
        for row_num in range(header_row + 1, header_row + len(analysis_results) + 1):
            self.ws.row_dimensions[row_num].height = 60

    def save(self):
        """Save the Excel file"""
        try:
            self.wb.save(self.output_filename)
            print(f"\nExcel report saved successfully: {self.output_filename}")
            return True
        except Exception as e:
            print(f"\nError saving Excel file: {str(e)}")
            return False

    def add_error_sheet(self, errors: List[Dict]):
        """
        Add a sheet with processing errors

        Args:
            errors: List of error dictionaries
        """
        if not errors:
            return

        # Create new sheet
        error_ws = self.wb.create_sheet("Errors")

        # Add headers
        error_ws['A1'] = "Files with Processing Errors"
        error_ws['A1'].font = Font(size=14, bold=True)
        error_ws.merge_cells('A1:C1')

        headers = ['Filename', 'Error Type', 'Error Message']
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = error_ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Add error data
        for idx, error in enumerate(errors, 1):
            row = 3 + idx
            error_ws.cell(row=row, column=1, value=error.get('filename', ''))
            error_ws.cell(row=row, column=2, value=error.get('error_type', ''))
            error_ws.cell(row=row, column=3, value=error.get('error_message', ''))

        # Adjust column widths
        error_ws.column_dimensions['A'].width = 30
        error_ws.column_dimensions['B'].width = 20
        error_ws.column_dimensions['C'].width = 60
